# core_final_processor_complete.py
# 完全對齊原始 GUI 程式邏輯與 Excel 輸出格式

import re
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
import xml.etree.ElementTree as ET
from openpyxl.styles import PatternFill, Font, Alignment
import os
import glob
import shutil
import subprocess
from datetime import datetime, timedelta
import time
from typing import Dict, Optional, List, Tuple


_IMPORTANT_EV = {
    # 避免與字幕底色衝突(FFFFCC/D4E6B7/FFDAB9/DDEBF7)，改用柔和色系
    # -------- 播出控制/節目單異動 (核心操作) --------
    '開始播出':                        ('🟢 開始播出',      'E8F5E9'),  # 淡綠
    '停止播出':                        ('🛑 停止播出',      'FFEBEE'),  # 淡紅
    '新增節目單':                      ('📋 新增節目單',   'FFF9F2'),  # 極淡珊瑚
    '發送節目單':                      ('📋 發送節目單',   'FFF4EB'),  # 淡珊瑚
    '發送的目前單':                    ('📋 發送節目單',   'FFF4EB'),
    'Recv HeartBeat Update Play List Msg': ('📋 更新播放列表', 'FFF4EB'),
    'Replace Playing List Old Start': ('⛔ 替換播出單(舊)', 'FFF0ED'),  # 柔和粉
    'Replace Playing List New Start': ('✅ 替換播出單(新)', 'F1F8E9'),  # 柔和綠
    
    # -------- 素材/編輯操作 ------------------
    '往播出列表拖入素材':              ('📥 拖入素材',      'E1F5FE'),  # 淡藍
    '貼上節目':                        ('📌 貼上節目',      'E8F5E9'),  # 淡綠
    '播出刪除節目':                    ('🗑 刪除節目',      'FFEBEE'),  # 淡紅
    '編輯當前節目':                    ('📝 編輯節目',      'FFF3E0'),  # 淡橘
    '刷新素材信息':                    ('🔄 刷新素材',      'F5F5F5'),  # 淺灰
    '服務通知素材刷新事件':            ('🔄 服務通知刷新',   'F5F5F5'),
    '素材檢查':                        ('🔍 素材檢查',      'E1F5FE'),  # 淡藍
    '素材檢查發現錯誤':                ('⚠ 素材錯誤',       'FFEBEE'),  # 淡紅
    
    # -------- 系統控制/警示 ------------------
    '向監控發送未上載條目報警信息':    ('⚠️ 未上載報警',   'FFF3E0'),  # 淡橘
    '完全接管':                        ('🔴 完全接管',      'FFEBEE'),  # 淡紅
    '鎖定':                            ('🔒 鎖定',          'E3F2FD'),  # 淡藍
    '解鎖':                            ('🔓 解鎖',          'E3F2FD'),
}

# 角色識別關鍵字（從 Ev_Info 中搜尋）
_ROLE_KEYWORDS = {'主機': '主機', '備機': '備機', 'master': '主機', 'backup': '備機'}

_DXAPLAYER_RE = re.compile(
    r'^(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})'    # 時間戳 (秒)
    r'[^\]]+\]'                                        # [...] 部分
    r'\s+Ev_Name=(.+?)\s+Ev_Info='                     # 尋找 Ev_Name 直到 Ev_Info=
    r'(.+?)\s+Ev_Hint='                                # Ev_Info 直到 Ev_Hint=
)


def _clean_ev_info(ev_info: str) -> str:
    """
    簡化 Ev_Info：移除 UUID/频道前綴/聲道譛註/標籤，保留節目名稱和重要資訊。
    範例:
      '節目名:MUCH-帝女花{447,...},{...}' -> 'MUCH-帝女花'
      '拖入素材名:MUCH-帝女花{447,...},{...}' -> 'MUCH-帝女花'
      '年代MUCH-主機, Start Open' -> 'Start Open'
    """
    import re as _re
    s = ev_info.strip()
    # 1. 移除頻道/機器前綴（字元-主機, / 字元-備機, 等）
    s = _re.sub(r'^.{2,15}(?:-主機|-備機|_master|_backup),\s*', '', s)
    # 2. 移除標籤（節目名:/拖入素材名:）
    s = _re.sub(r'^節目名:', '', s)
    s = _re.sub(r'^拖入素材名:', '', s)
    # 3. 移除 {UUID,...} 區塊（UUID 模式）
    s = _re.sub(r'\{[0-9a-fA-F]{8}-[^\}]*\}', '', s)
    # 4. 移除 {xxx:...} 堀位（{聲道:...} {频道:...} 等）
    s = _re.sub(r',?\s*\{[^\}\{]{0,12}:[^\}]*\}', '', s)
    # 5. 移除 ,{N, xxx-xxx-xxx, ...} 累購識別區塊
    s = _re.sub(r',?\s*\{\d+,\s*[0-9a-fA-F\-]{8,}[^\}]*\}', '', s)
    # 6. 特殊處理：刪除節目 XML 格式 (strClipFile="..." strClipName="...")
    if 'strClipFile=' in s or 'strClipName=' in s:
        file_m = _re.search(r'strClipFile="([^"]+)"', s)
        name_m = _re.search(r'strClipName="([^"]+)"', s)
        parts = []
        if file_m: parts.append(f'strClipFile="{file_m.group(1)}"')
        if name_m: parts.append(f'strClipName="{name_m.group(1)}"')
        if parts:
            s = " ".join(parts) + " />"
    # 7. 清除多餘逗號/空白
    s = _re.sub(r'[,\s]{2,}', ' ', s.strip())
    s = _re.sub(r'^[,\s]+', '', s)
    s = _re.sub(r'[,\s]+$', '', s)
    # 7. 移除 VIPS 錯誤詳情（--<VIPS-...:...>）
    s = _re.sub(r'--<[^>]*>', '', s)
    s = _re.sub(r',?\s*<[A-Z][A-Z0-9\-]*:[^>]*>', '', s)
    # 8. 最終整理
    s = _re.sub(r'[,\s]{2,}', ' ', s.strip())
    s = _re.sub(r'^[,\s]+', '', s)
    s = _re.sub(r'[,\s]+$', '', s)
    return s.strip() or ev_info[:60]

def process_dxaplayer_events(config: Dict, target_date: datetime, callback) -> list:
    """
    解析主備兩台的 DxAplayerLog，回傳重點事件 list of dict：
      _ts_dt, 時間(HH:MM:SS), 機器, 角色, 事件類型, 事件詳情, ev_color
    不修改任何 DataFrame，呼叫端自行決定如何使用。
    """
    target_str = target_date.strftime('%Y%m%d')
    log_filename = f'DxAplayerLog{target_str}.log'
    events = []

    for base_path in config.get('log_base_paths', []):
        m_ip = re.search(r'\\\\([\d.]+)\\', base_path)
        machine_ip = m_ip.group(1) if m_ip else base_path
        
        # --- 多層級路徑探測 (解決 151 C$ 被拒問題) ---
        possible_paths = [base_path]
        if 'c$' in base_path.lower():
            # 嘗試直接共享 (例如 \\172.25.80.151\dayang\...)
            alt = base_path.replace(r'\c$\dayang', r'\dayang').replace(r'\C$\dayang', r'\dayang')
            if alt not in possible_paths: possible_paths.append(alt)
            # 嘗試 D$
            alt_d = base_path.replace(r'\c$', r'\d$').replace(r'\C$', r'\D$')
            if alt_d not in possible_paths: possible_paths.append(alt_d)

        log_path = None
        for p_cand in possible_paths:
            full_cand = os.path.join(p_cand, log_filename)
            if os.path.exists(full_cand):
                log_path = full_cand
                callback(f'DxAplayerLog 路徑探測成功: {log_path}')
                break
        
        if not log_path:
            callback(f'DxAplayerLog 均不可存取: {machine_ip} (嘗試過 {len(possible_paths)} 種路徑)')
            continue

        m_ip = re.search(r'\\\\([\d.]+)\\', log_path)
        machine_ip = m_ip.group(1) if m_ip else base_path

        lines = None
        for enc in ('cp950', 'big5hkscs', 'gbk', 'utf-8', 'latin1'):
            try:
                with open(log_path, 'r', encoding=enc, errors='replace') as f:
                    lines = f.readlines()
                callback(f'DxAplayerLog 讀取成功({enc}): {machine_ip}')
                break
            except Exception:
                continue
        if lines is None:
            callback(f'DxAplayerLog 讀取失敗: {machine_ip}', 'error')
            continue

        found = 0
        for line in lines:
            m = _DXAPLAYER_RE.match(line)
            if not m:
                continue
            ts_str, ev_name, ev_info = m.group(1), m.group(2).strip(), m.group(3).strip()

            # 優先權 1：精確全名比對 (不分中英文)
            matched_key = None
            if ev_name in _IMPORTANT_EV:
                matched_key = ev_name
            
            # 優先權 2：如果您定義的 key 是英文 (ASCII)，則不支援部分匹配，必須全名一致
            if matched_key is None:
                for k in _IMPORTANT_EV:
                    if k.isascii():
                        if ev_name == k:
                            matched_key = k
                            break
                    else:
                        # 中文：支援包含比對或起頭比對 (針對亂碼環境)
                        if k in ev_name or ev_name.startswith(k[:2]):
                            matched_key = k
                            break
            if matched_key is None:
                continue

            # 從 Ev_Info 識別角色（主機/備機），否則用 IP 區分
            role = '-'
            for kw, role_label in _ROLE_KEYWORDS.items():
                if kw in ev_info:
                    role = role_label
                    break
            if role == '-':
                # 從 Ev_Info 節目單資訊推斷（如「年代MUCH-主機」「年代MUCH-備機」）
                if any(x in ev_info for x in ('主機', '-master', 'Master')):
                    role = '主機'
                elif any(x in ev_info for x in ('備機', '-backup', 'Backup')):
                    role = '備機'
            try:
                ts_dt = datetime.strptime(ts_str, '%Y-%m-%d %H:%M:%S')
            except ValueError:
                ts_dt = None

            label, color = _IMPORTANT_EV[matched_key]
            events.append({
                '_ts_dt':   ts_dt,
                '時間':     ts_str[11:19],
                '機器':     machine_ip,
                '角色':     role,
                '事件類型': label,
                '事件詳情': ev_info,
                'ev_color': color,
            })
            found += 1

        callback(f'DxAplayerLog 重點事件: {found} 筆 ({machine_ip})')

    events.sort(key=lambda e: e['_ts_dt'] or datetime.min)
    return events


# ==================== 頻道配置 ====================
def get_channel_config(channel: str, target_date: datetime) -> Dict:
    target_str_yyyymmdd = target_date.strftime("%Y%m%d")
    target_str_yymmdd = target_date.strftime("%y%m%d")
    target_str_date = target_date.strftime("%Y-%m-%d")

    if channel == "CH04":
        channel_display = '年代MUCH'
        final_output_prefix = 'CH04年代MUCH台'
        lis_prefix = '年代MUCH'
        lis_old_suffix = '080350'
        network_channels = ["3"]
        output_prefix = '年代MUCH播出紀錄'
    elif channel == "CH05":
        channel_display = '壹綜合'
        final_output_prefix = 'CH05壹綜合台'
        lis_prefix = '壹綜合'
        lis_old_suffix = '061744' # 預設時間
        network_channels = ["1"] # 綜合台如果只有一軌那就預設 1，先放 ["1"] 或 ["3", "4"] 依據實際。但暫不確定就預設 ["3"] 或 ["4"]，我們延用 CH02 的 ["3", "4"] 避免遺漏。
        network_channels = ["3", "4"]
        output_prefix = '壹綜合播出紀錄'
    else:  # CH02 Default
        channel_display = '壹電影'
        final_output_prefix = 'CH02壹電影台'
        lis_prefix = '壹電影'
        lis_old_suffix = '061744'
        network_channels = ["3", "4"]
        output_prefix = '壹電影播出紀錄'

    config = {
        'channel': channel,
        'channel_display': channel_display,
        'final_output_prefix': final_output_prefix,
        'lis_pattern': f"{lis_prefix}-{target_str_yymmdd}*.lis",
        'lis_old_file': f"{lis_prefix}-{target_str_yymmdd}-{lis_old_suffix}old.lis",
        'log_patterns': [
            f"XBlog[{channel}--主機]{target_str_yyyymmdd}.txt",
            f"XBlog[{channel}--備機]{target_str_yyyymmdd}.txt",
            f"XBlog{target_str_yyyymmdd}.txt"
        ],
        'subtitle_file': f"{channel} - 字幕播出單_{{{target_str_date}}}.log",
        'network_channels': network_channels,
        'output_prefix': output_prefix,
    }

    if channel == "CH04":
        config.update({
            'lis_base_paths': [
                r'\\172.25.80.141\c$\dayang\obin_26011601_2.8_NIANDAI\autobak',
                r'\\172.25.80.142\c$\dayang\obin_26011601_2.8_NIANDAI\autobak'
            ],
            'log_base_paths': [
                r'\\172.25.80.141\c$\dayang\obin_26011601_2.8_NIANDAI\log',
                r'\\172.25.80.142\c$\dayang\obin_26011601_2.8_NIANDAI\log'
            ],
            'subtitle_base_path': r'\\172.25.80.34\d$\emagic3_5_patch_2025-4-41_test\Other',
        })
    elif channel == "CH05":
        config.update({
            'lis_base_paths': [
                r'\\172.25.80.151\c$\dayang\obin_26011601_2.8_NIANDAI\autobak',
                r'\\172.25.80.152\c$\dayang\obin_26011601_2.8_NIANDAI\autobak'
            ],
            'log_base_paths': [
                r'\\172.25.80.151\c$\dayang\obin_26011601_2.8_NIANDAI\log',
                r'\\172.25.80.152\c$\dayang\obin_26011601_2.8_NIANDAI\log'
            ],
            'subtitle_base_path': r'\\172.25.80.35\d$\emagic3_5_patch_2025-4-41_test\Other',
        })
    else:  # CH02 Default
        config.update({
            'lis_base_paths': [
                r'\\172.25.80.121\c$\dayang\obin_26011601_2.8_NIANDAI\autobak',
                r'\\172.25.80.122\c$\dayang\obin_26011601_2.8_NIANDAI\autobak'
            ],
            'log_base_paths': [
                r'\\172.25.80.121\c$\dayang\obin_26011601_2.8_NIANDAI\log',
                r'\\172.25.80.122\c$\dayang\obin_26011601_2.8_NIANDAI\log'
            ],
            'subtitle_base_path': r'\\172.25.80.32\d$\emagic3_5_patch_2025-4-41_test\Other',
        })

    return config


# ==================== 時間碼轉換 ====================
def frames_to_drop_frame_timecode(total_frames: int) -> str:
    """轉換幀數為 29.97 fps drop-frame timecode (HH:MM:SS:FF)"""
    try:
        total_frames = int(total_frames)
        if total_frames < 0:
            return "00:00:00:00"

        frames_in_10_minutes = 17982
        frames_in_1_minute = 1798
        ten_minutes_count = total_frames // frames_in_10_minutes
        frames_remainder = total_frames % frames_in_10_minutes
        one_minute_count = (frames_remainder - 2) // frames_in_1_minute if frames_remainder > 2 else 0
        total_drops = (ten_minutes_count * 9 + one_minute_count) * 2
        total_non_drop_frames = total_frames + total_drops

        framerate = 30
        ff = total_non_drop_frames % framerate
        ss = (total_non_drop_frames // framerate) % 60
        mm = (total_non_drop_frames // (framerate * 60)) % 60
        hh = total_non_drop_frames // (framerate * 3600)

        return f"{hh:02d}:{mm:02d}:{ss:02d}:{ff:02d}"
    except (ValueError, TypeError):
        return "00:00:00:00"


def frames_to_timecode(total_frames, fps=30):
    """將總幀數轉換為時分秒幀格式 HH:MM:SS:FF"""
    if total_frames is None or total_frames == 0:
        return "00:00:00:00"

    frames = int(total_frames % fps)
    total_seconds = int(total_frames // fps)

    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    seconds = total_seconds % 60

    return f"{hours:02d}:{minutes:02d}:{seconds:02d}:{frames:02d}"


# ==================== LIS 處理（完整版本 - 性能優化）====================
def process_lis_file_complete(config: Dict, target_date: datetime, callback) -> Tuple[pd.DataFrame, List]:
    """完整的 LIS 處理邏輯，包含去重與優先級處理 - 優化版本"""
    target_str_yyyymmdd = target_date.strftime("%Y%m%d")
    target_str_yymmdd = target_date.strftime("%y%m%d")
    yesterday_str_yymmdd = (target_date - timedelta(days=1)).strftime("%y%m%d")

    search_paths = config.get('lis_base_paths', [config.get('lis_base_path')])
    unique_items: Dict[str, Tuple[ET.Element, int]] = {}
    files_to_process = []

    callback(f"開始處理 {config['channel_display']} {target_date.strftime('%Y年%m月%d日')} 的播出記錄...")

    try:
        for search_path in search_paths:
            # 依據頻道決定字首與特定時間的 old.lis 檔名
            if config['channel'] == 'CH04':
                prefix = '年代MUCH'
                old_suffix = '080350'
            elif config['channel'] == 'CH05':
                prefix = '壹綜合'
                old_suffix = '061744'
            else:
                prefix = '壹電影'
                old_suffix = '061744'

            # 多層級路徑試探 (解決 151 C$ 被拒問題)
            possible_paths = [search_path]
            if 'c$' in search_path.lower():
                alt = search_path.replace(r'\c$\dayang', r'\dayang').replace(r'\C$\dayang', r'\dayang')
                if alt not in possible_paths: possible_paths.append(alt)
                alt_d = search_path.replace(r'\c$', r'\d$').replace(r'\C$', r'\D$')
                if alt_d not in possible_paths: possible_paths.append(alt_d)

            actual_base = None
            for p_cand in possible_paths:
                if os.path.exists(p_cand):
                    actual_base = p_cand
                    break
            
            if not actual_base:
                callback(f"路徑不可存取: {search_path}", 'warning')
                continue

            old_file_path = os.path.join(actual_base, config['lis_old_file'])
            yesterday_pattern = os.path.join(actual_base, f"{prefix}-{yesterday_str_yymmdd}*.lis")
            yesterday_old_file = os.path.join(actual_base, f"{prefix}-{yesterday_str_yymmdd}-{old_suffix}old.lis")

            try:
                # 收集前一天的檔案
                if os.path.exists(yesterday_old_file):
                    files_to_process.append((yesterday_old_file, "old", "yesterday"))

                yesterday_files = glob.glob(yesterday_pattern)
                for file_path in yesterday_files:
                    if not file_path.endswith("old.lis"):
                        files_to_process.append((file_path, "latest", "yesterday"))

                # 收集當天的檔案
                if os.path.exists(old_file_path):
                    files_to_process.append((old_file_path, "old", "today"))

                today_pattern = os.path.join(actual_base, config['lis_pattern'])
                today_files = glob.glob(today_pattern)
                for file_path in today_files:
                    if not file_path.endswith("old.lis"):
                        files_to_process.append((file_path, "latest", "today"))
            except Exception as e:
                callback(f"搜尋檔案錯誤 {actual_base}: {e}", 'error')

        if not files_to_process:
            callback(f"找不到 {config['channel_display']} 相關的 .lis 檔案", 'error')
            callback(f"搜尋路徑與模式如下：", 'error')
            for sp in search_paths:
                callback(f"  - 路徑: {sp}", 'error')
                callback(f"  - 模式: {config['lis_pattern']}", 'error')
            return pd.DataFrame(), []

        callback(f"共找到 {len(files_to_process)} 個相關檔案")
        initial_record_count = 0

        # 處理所有檔案
        for file_path, source_type, file_date in files_to_process:
            try:
                callback(f"正在讀取檔案: {file_path}")
                tree = ET.parse(file_path)
                items = tree.getroot().findall(".//SItem")

                for item in items:
                    rptm = item.get("RPTM", "")

                    # 快速過濾：只檢查日期部分
                    if not rptm or len(rptm) < 8:
                        continue

                    date_part = rptm[:8].replace(' ', '')

                    # 判斷是否應包含此記錄
                    include_record = False
                    if file_date == "yesterday":
                        if date_part == target_str_yyyymmdd:
                            include_record = True
                        elif len(rptm) >= 19:
                            try:
                                hour = int(rptm[9:11])
                                minute = int(rptm[12:14])
                                if hour == 23 and minute >= 45:
                                    include_record = True
                            except (ValueError, IndexError):
                                pass
                    elif file_date == "today":
                        if date_part == target_str_yyyymmdd:
                            include_record = True

                    if not include_record:
                        continue

                    initial_record_count += 1

                    # 使用穩定的鍵進行去重
                    clip_f = item.get("ClipF", "")
                    ctid = item.get("CTID", "")
                    tin = item.get("TIn", "")
                    key = f"{rptm}_{clip_f}_{ctid}_{tin}"

                    # 優先級：latest > old, today > yesterday
                    source_priority = 2 if source_type == "latest" else 1
                    date_priority = 2 if file_date == "today" else 1
                    total_priority = source_priority + date_priority

                    if key in unique_items:
                        _old_item, existing_priority = unique_items[key]
                        if total_priority > existing_priority:
                            unique_items[key] = (item, total_priority)
                    else:
                        unique_items[key] = (item, total_priority)

            except Exception as e:
                callback(f"處理檔案錯誤 {os.path.basename(file_path)}: {e}", 'error')
                continue

        # 取得最終項目並排序
        final_items = [v[0] for v in unique_items.values()]
        final_items.sort(key=lambda x: x.get("RPTM", ""))

        callback(
            f"去重複處理完成 - 原始: {initial_record_count}, 最終: {len(final_items)}, 移除: {initial_record_count - len(final_items)}")

        if not final_items:
            callback(f"沒有找到 {config['channel_display']} 的有效播出記錄", 'error')
            return pd.DataFrame(), []

        # 批量轉換為 DataFrame（避免逐行 append）
        ctid_mapping = {"001": "PROG", "002": "PROMO", "AD": "AD", "003": "AD"}
        data = []

        for sitem in final_items:
            name = sitem.get("Name", "")
            ctid = sitem.get("CTID", "")
            mapped_ctid = ctid_mapping.get(ctid, ctid)

            # 節目類型判斷
            if name.startswith('節目'):
                mapped_ctid = 'LIVE'

            rptm = sitem.get("RPTM", "")
            segment = sitem.get("SegIndex", "") or ""
            clip_f = sitem.get("ClipF", "")

            # 預先轉換時間碼
            tin_tc = frames_to_drop_frame_timecode(sitem.get("TIn", ""))
            dura_tc = frames_to_drop_frame_timecode(sitem.get("Dura", ""))
            actual_tc = frames_to_drop_frame_timecode(sitem.get("ActualDura", ""))

            # 處理字幕
            cg_items = sitem.findall("CGtwnd/CGItem")

            base_row = {
                '表定播放時間 (RPTM)': rptm,
                '節目名 (Name)': name,
                '段': segment,
                '入點 (TIn)': tin_tc,
                '持續時間 (Dura)': dura_tc,
                '實際長度': actual_tc,
                '影片檔名': clip_f,
                '類型': mapped_ctid,
            }

            if not cg_items:
                base_row['字幕指令'] = ""
                base_row['字幕Dur'] = ""
                data.append(base_row)
            else:
                # 字幕排序：G > BLGL/BLGLM > others
                g_items = [cg for cg in cg_items if cg.get("KeyName", "") == "G"]
                blgl_items = [cg for cg in cg_items if cg.get("KeyName", "") in ["BLGLM", "BLGL"]]
                other_items = [cg for cg in cg_items if cg.get("KeyName", "") not in ["G", "BLGLM", "BLGL"]]
                final_order = g_items + blgl_items + other_items

                for cg in final_order:
                    key_name = cg.get("KeyName", "")
                    if key_name:
                        row = base_row.copy()
                        row['字幕指令'] = key_name
                        row['字幕Dur'] = frames_to_drop_frame_timecode(cg.get("Dur", ""))
                        data.append(row)

        # 一次性建立 DataFrame
        df = pd.DataFrame(data)
        callback(f"LIS 處理完成，共 {len(df)} 筆記錄")
        return df, final_items

    except Exception as e:
        callback(f"LIS處理錯誤: {e}", 'error')
        return pd.DataFrame(), []


# ==================== 網路日誌處理（完整版本）====================
def process_network_log_complete(config: Dict, target_date: datetime, callback) -> Optional[pd.DataFrame]:
    """完整的網路日誌處理，包含 Cue/Play 配對"""
    target_str_yyyymmdd = target_date.strftime('%Y%m%d')

    log_paths_base = config.get('log_base_paths', [config.get('log_base_path')])
    log_file_patterns = config.get('log_patterns', [config.get('log_file')])
    log_file_paths = []
    
    # 組合所有可能的路徑與檔名
    for base in log_paths_base:
        for pattern in log_file_patterns:
            log_file_paths.append(os.path.join(base, pattern))

    callback(f"搜索 {config['channel_display']} 網路日誌檔案...")

    # 放寬規則：支持兩種格式
    # 舊版：12-28 11:52:40:893 [YY-dd hh:mm:ss:sss]<3,VS.Cue
    # 新版：2026-01-28 11:52:40.893 [C= 1 ...]<3,VS.Cue
    log_pattern = re.compile(
        r"(\d{2,4}-\d{2}-\d{2}|\d{2}-\d{2})\s+(\d{2}:\d{2}:\d{2}[.:]\d{3}).*?<(\d+),VS\.(Cue|Play)(?:<([^,>]+),.*?>)?"
    )

    data = []
    current_cue = None
    found_log_file = False

    for log_path in log_file_paths:
        try:
            if not os.path.exists(log_path):
                callback(f"檔案不存在: {log_path}", 'warning')
                continue

            callback(f"找到網路日誌檔案: {log_path}")

            # 嘗試多種編碼
            encodings = ['cp1252', 'utf-8', 'latin1']
            file_content = None
            for encoding in encodings:
                try:
                    with open(log_path, 'r', encoding=encoding) as f:
                        file_content = f.readlines()
                        found_log_file = True
                        break
                except UnicodeDecodeError:
                    continue

            if file_content is None:
                callback(f"無法讀取檔案: {log_path}", 'error')
                continue

            processed_lines = 0
            match_count = 0
            for line in file_content:
                match = log_pattern.search(line)
                if match:
                    # 擷取日期 (處理 YYYY-MM-DD -> MM-DD)
                    date_val = match.group(1)
                    if len(date_val) > 5: # 2026-01-28
                        date_str = date_val[-5:] # 01-28
                    else:
                        date_str = date_val
                        
                    time_str = match.group(2).replace('.', ':') # 統一轉為 : 格式
                    channel = match.group(3)
                    event_type = match.group(4)
                    filename = match.group(5)

                    target_channels = config['network_channels']

                    if channel in target_channels:
                        full_time_str = f"{target_date.year}-{date_str} {time_str}"
                        processed_lines += 1
                        match_count += 1

                        if event_type == "Cue" and filename:
                            current_cue = {"CueDown": full_time_str, "節目檔名": filename}
                        elif event_type == "Play" and current_cue:
                            current_cue["實際播出"] = full_time_str
                            callback(f"匹配播出記錄: {current_cue['節目檔名']} 實際播出={full_time_str}")
                            data.append(current_cue)
                            current_cue = None

            callback(f"處理了 {processed_lines} 行網路日誌記錄")
            
            # 如果還是找不到，進行關鍵字探測
            if match_count == 0:
                callback(f"⚠️ 找不到正規化記錄，開始關鍵字探測...", 'warning')
                keywords = ["VS.Cue", "VS.Play", "Cue", "Play"]
                found_samples = []
                for line in file_content:
                    for kw in keywords:
                        if kw in line:
                            found_samples.append(line.strip()[:150])
                            break
                    if len(found_samples) >= 3:
                        break
                
                if found_samples:
                    callback(f"在此檔案中發現以下疑似記錄：", 'info')
                    for s in found_samples:
                        callback(f"  - {s}", 'info')
                else:
                    callback(f"在此檔案中未發現任何 VS.Cue 或 Cue 關鍵字", 'warning')

            if len(data) > 0:
                break  # 找到並有有效記錄才退出
            else:
                callback(f"檔案 {log_path} 內無匹配頻道記錄，嘗試搜尋下一個位置...", 'warning')

        except Exception as e:
            callback(f"處理 {log_path} 錯誤: {e}", 'error')

    if not found_log_file:
        callback(f"找不到任何網路日誌檔案", 'error')
        callback("嘗試搜尋的完整路徑清單如下：", 'warning')
        for lp in log_file_paths:
            callback(f"  - {lp}", 'warning')
        return None

    if not data:
        callback(f"在網路日誌中找不到相關記錄", 'error')
        return None

    df = pd.DataFrame(data)[["CueDown", "實際播出", "節目檔名"]]
    callback(f"網路日誌處理完成: {len(df)} 筆記錄")
    return df


# ==================== 字幕網絡連線 ====================
def try_subtitle_network_connection(server_ip: str, callback):
    """嘗試建立字幕機的網路連線"""
    username = r"Administrator"
    password = "Abc1234"
    
    # 清理舊連線
    try:
        subprocess.run(
            f'net use "\\\\{server_ip}\\d$" /delete /yes',
            shell=True,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            timeout=5
        )
    except:
        pass
        
    methods = [
        f'net use "\\\\{server_ip}\\d$" "{password}" /user:"{username}"',
        f'net use "\\\\{server_ip}\\d$" "{password}"',
        f'net use "\\\\{server_ip}\\d$" "{password}" /user:"{server_ip}\\pres"',
    ]

    for method_idx, cmd in enumerate(methods, 1):
        try:
            subprocess.run(
                cmd,
                shell=True,
                check=True,
                capture_output=True,
                text=True,
                timeout=15
            )
            callback(f"字幕機連線成功：{server_ip} (方法 {method_idx})", 'info')
            return True
        except:
            continue

    # 檢查是否已連接
    try:
        check_cmd = f'net use "\\\\{server_ip}\\d$"'
        result = subprocess.run(
            check_cmd,
            shell=True,
            check=True,
            capture_output=True,
            text=True,
            timeout=5
        )
        if "OK" in result.stdout or "成功" in result.stdout:
            callback(f"檢測到已存在的字幕機連線：{server_ip}", 'info')
            return True
    except:
        pass

    return False

# ==================== 字幕日誌處理（完整版本）====================
def process_subtitle_log_complete(config: Dict, target_date: datetime, callback) -> Optional[pd.DataFrame]:
    """完整的字幕日誌處理"""
    target_str = target_date.strftime('%Y-%m-%d')

    server_path = config['subtitle_base_path']
    
    # 提取 IP 地址
    ip_match = re.search(r'\\\\(\d+\.\d+\.\d+\.\d+)', server_path)
    if ip_match:
        server_ip = ip_match.group(1)
        callback(f"嘗試連線至字幕機: {server_ip}...")
        try_subtitle_network_connection(server_ip, callback)

    log_path = os.path.join(server_path, config['subtitle_file'])

    callback(f"搜索字幕日誌檔案...")

    if not os.path.exists(log_path):
        callback(f"找不到字幕日誌: {log_path}", 'info')
        return None

    try:
        with open(log_path, 'r', encoding='big5-hkscs') as file:
            log_content = file.read()
        callback(f"成功讀取字幕日誌檔案")
    except Exception as e:
        callback(f"讀取字幕日誌錯誤: {e}", 'error')
        return None

    lines = log_content.splitlines()

    # Pass 1: 收集 CUE 資訊
    cue_data = {}
    i = 0
    while i < len(lines):
        if 'Command:CUE.' in lines[i]:
            item_guid, item_name, total_frames = None, None, None
            for j in range(i, min(i + 10, len(lines))):
                guid_match = re.search(r'ItemGuid:([0-9A-F-]+)', lines[j])
                if guid_match:
                    item_guid = guid_match.group(1)

                name_match = re.search(r'ItemName:([^,]+)', lines[j])
                if name_match:
                    item_name = name_match.group(1).strip()

                frame_match = re.search(r'nTotalFrame:(\d+)', lines[j])
                if frame_match:
                    total_frames = int(frame_match.group(1))

                if item_guid and item_name and total_frames is not None:
                    if item_guid not in cue_data:
                        cue_data[item_guid] = {'name': item_name, 'frames': total_frames}
                        callback(f"[Pass 1] 找到 CUE: GUID={item_guid}, Name={item_name}, Frames={total_frames}")
                    break
        i += 1

    callback(f"收集到 {len(cue_data)} 筆 CUE 記錄")

    # Pass 2: 找 PLAYSTORY 並匹配
    play_events = []
    for line_num, line in enumerate(lines, 1):
        if 'Command:PLAYSTORY.' in line:
            callback(f"[Pass 2] 找到可能的 PLAYSTORY: {line.strip()[:60]}...")
            play_guid_match = re.search(r'PlayStoryList:\[ItemGuid:([0-9A-F-]+)\.StoryIndex:0\]', line)

            if not play_guid_match:
                continue

            play_guid = play_guid_match.group(1)

            if play_guid not in cue_data:
                continue

            timestamp_match = re.search(r'(\d{4}-\d{2}-\d{2}\s\d{2}:\d{2}:\d{2}:\d{4})', line)

            if not timestamp_match:
                continue

            full_timestamp = timestamp_match.group(1)
            event_date = full_timestamp[:10]

            if event_date != target_str:
                continue

            try:
                cue_info = cue_data[play_guid]
                item_name = cue_info['name']
                total_frames = cue_info['frames']

                datetime_part, millisec_part = full_timestamp.rsplit(':', 1)
                time_part = datetime_part.split(' ')[1]
                broadcast_time = f"{time_part}:{millisec_part}"
                duration_timecode = frames_to_timecode(total_frames)

                play_events.append({
                    '字幕指令': item_name,
                    '字幕Dur': duration_timecode,
                    '字幕實際播出時間': broadcast_time,
                    '原始時間戳': full_timestamp,
                    '總幀數': total_frames
                })
                callback(f"    -> 成功創建字幕事件: 指令={item_name}, 時間={broadcast_time}")
            except Exception as parse_error:
                callback(f"解析 PLAYSTORY 事件時出錯: {parse_error}", 'error')

    if not play_events:
        callback("字幕日誌中找不到有效 PLAYSTORY 事件", 'info')
        return None

    df = pd.DataFrame(play_events)

    # 去重
    num_before = len(df)
    if num_before > 0:
        df = df.drop_duplicates(subset=['字幕指令', '字幕實際播出時間', '字幕Dur'])
        num_after = len(df)
        if num_after < num_before:
            callback(f"移除了 {num_before - num_after} 筆重複的字幕記錄")

    # 建立 datetime 欄位用於排序
    df['字幕實際播出時間_dt'] = pd.to_datetime(df['原始時間戳'], format='%Y-%m-%d %H:%M:%S:%f', errors='coerce')
    df = df.sort_values('字幕實際播出時間_dt')

    callback(f"字幕日誌處理完成，共 {len(df)} 筆記錄")
    return df


# ==================== 時間欄位預處理 ====================
def preprocess_lis_time_col(col: pd.Series) -> pd.Series:
    """預處理 LIS 時間欄位"""
    col_fixed = col.str.replace(r'^(\d{4})(\d{2})(\d{2}) ', r'\1-\2-\3 ', regex=True)
    col_fixed = col_fixed.str.replace(r':(\d{2})$', r'.\1', regex=True)
    return pd.to_datetime(col_fixed, format='%Y-%m-%d %H:%M:%S.%f', errors='coerce')


def preprocess_log_time_col(col: pd.Series) -> pd.Series:
    """預處理 log 時間欄位"""
    col_fixed = col.str.replace(r':(\d{3})$', r'.\1', regex=True)
    return pd.to_datetime(col_fixed, format='%Y-%m-%d %H:%M:%S.%f', errors='coerce')


# ==================== Excel 生成（完整版本）====================
def generate_final_excel_complete(lis_df: pd.DataFrame, log_df: Optional[pd.DataFrame],
                                  subtitle_df: Optional[pd.DataFrame], config: Dict,
                                  target_date: datetime, output_path: str, callback,
                                  events: list = None):
    """完整的 Excel 生成邏輯，對齊原始程式的所有細節"""

    callback(f"開始合併與生成 {config['channel_display']} 報表...")

    target_str_yyyymmdd = target_date.strftime("%Y%m%d")

    try:
        # ========== 準備網路日誌 ==========
        if log_df is not None and not log_df.empty:
            log_df.columns = log_df.columns.str.strip()
            log_df['節目檔名'] = log_df['節目檔名'].astype(str).str.strip()
            log_df['播出時間_dt'] = preprocess_log_time_col(log_df['實際播出'])

        # ========== 準備 LIS 資料 ==========
        lis_df.columns = lis_df.columns.str.strip()
        lis_df['影片檔名'] = lis_df['影片檔名'].astype(str).str.strip()
        lis_df['表定播放時間_dt'] = preprocess_lis_time_col(lis_df['表定播放時間 (RPTM)'])

        # ========== 合併 LIS 與網路日誌 ==========
        if log_df is not None and not log_df.empty:
            log_df = log_df.rename(columns={"節目檔名": "影片檔名"})
            merged = pd.merge(lis_df, log_df, how='left', on='影片檔名', suffixes=('', '_log'))
            merged['時間差'] = (merged['播出時間_dt'] - merged['表定播放時間_dt']).dt.total_seconds().abs().fillna(0)

            # 時間差過濾 (<=1秒)
            filtered = merged[merged['時間差'] <= 1].copy()
            callback(f"時間差過濾後 (<=1s)，剩下 {len(filtered)} 筆記錄")
        else:
            filtered = lis_df.copy()
            filtered['CueDown'] = ''
            filtered['實際播出'] = ''
            filtered['播出時間_dt'] = pd.NaT
            filtered['時間差'] = 0
            callback("沒有網路日誌，使用 LIS 資料")

        # ========== 分離 LIVE 節目以避免錯誤去重 ==========
        callback("分離 LIVE 節目以避免錯誤去重...")
        filtered = filtered.sort_values('時間差', ascending=True)

        live_mask = filtered['類型'] == 'LIVE'
        live_programs = filtered[live_mask]
        other_programs = filtered[~live_mask]
        callback(f"找到 {len(live_programs)} 筆 LIVE 節目，{len(other_programs)} 筆其他節目。")

        # 對非 LIVE 節目去重
        dedup_subset = ['影片檔名', '實際播出']
        before_dedup_count = len(other_programs)
        other_programs_deduped = other_programs.drop_duplicates(subset=dedup_subset, keep='first')
        after_dedup_count = len(other_programs_deduped)

        if before_dedup_count > after_dedup_count:
            callback(
                f"對非 LIVE 節目移除重複的播出匹配後，剩下 {after_dedup_count} 筆記錄 (移除了 {before_dedup_count - after_dedup_count} 筆)")

        # 合併回來
        filtered = pd.concat([live_programs, other_programs_deduped], ignore_index=True)
        filtered = filtered.sort_values(by='表定播放時間_dt').reset_index(drop=True)
        callback(f"合併後總記錄數: {len(filtered)}")

        # ========== 字幕處理邏輯 ==========
        programs_df = filtered.copy()
        programs_df['prog_start_time'] = programs_df['播出時間_dt'].fillna(programs_df['表定播放時間_dt'])
        programs_df = programs_df.dropna(subset=['prog_start_time']).sort_values('prog_start_time')

        all_subs_df = subtitle_df.copy() if subtitle_df is not None and not subtitle_df.empty else pd.DataFrame()
        new_subtitle_rows = []

        if not all_subs_df.empty and not programs_df.empty:
            callback(f"開始匹配 {len(all_subs_df)} 筆字幕記錄到節目排程...")

            subs_with_prog = pd.merge_asof(
                all_subs_df.sort_values('字幕實際播出時間_dt'),
                programs_df,
                left_on='字幕實際播出時間_dt',
                right_on='prog_start_time',
                direction='nearest',
                tolerance=pd.Timedelta('1m')
            )

            for _, row in subs_with_prog.iterrows():
                try:
                    if pd.notna(row['prog_start_time']):
                        sort_time = row['prog_start_time'] + pd.Timedelta('1ms')
                        new_row = {
                            '節目名 (Name)': row['節目名 (Name)'],
                            '段': row['段'],
                            '影片檔名': row['影片檔名'],
                            '類型': '字幕',
                            '字幕指令': row['字幕指令_x'],
                            '字幕Dur': row['字幕Dur_x'],
                            '字幕實際播出時間': row['字幕實際播出時間'],
                            '排序時間': sort_time,
                            '表定播放時間 (RPTM)': '',
                            'CueDown': '',
                            '實際播出': '',
                            '時間差': 0,
                            '入點 (TIn)': '',
                            '實際長度': ''
                        }
                        new_subtitle_rows.append(new_row)
                    else:
                        # 主控添加
                        new_subtitle_rows.append({
                            '類型': '主控添加',
                            '字幕指令': row['字幕指令_x'],
                            '字幕Dur': row['字幕Dur_x'],
                            '字幕實際播出時間': row['字幕實際播出時間'],
                            '排序時間': row['字幕實際播出時間_dt'],
                            '節目名 (Name)': '', '段': '', '影片檔名': '',
                            '表定播放時間 (RPTM)': '', 'CueDown': '', '實際播出': '',
                            '時間差': 0, '入點 (TIn)': '', '實際長度': ''
                        })
                except KeyError as e:
                    callback(f"處理字幕時發生 KeyError: {e}", 'error')
                    continue

        # ========== 整合字幕到主表 ==========
        filtered['排序時間'] = filtered['播出時間_dt'].fillna(filtered['表定播放時間_dt'])
        filtered['字幕指令'] = ''
        filtered['字幕Dur'] = ''
        filtered['字幕實際播出時間'] = ''

        if new_subtitle_rows:
            new_subs_df = pd.DataFrame(new_subtitle_rows)
            filtered = pd.concat([filtered, new_subs_df], ignore_index=True)

        filtered = filtered.sort_values(by=['排序時間', '字幕實際播出時間']).reset_index(drop=True)
        filtered = filtered.drop('排序時間', axis=1)

        # ========== 最終過濾 ==========
        before_final_filter = len(filtered)
        final_filter_condition = (
                (filtered['CueDown'].notna() & (filtered['CueDown'] != '')) |
                (filtered['實際播出'].notna() & (filtered['實際播出'] != '')) |
                (filtered['字幕實際播出時間'].notna() & (filtered['字幕實際播出時間'] != '')) |
                (filtered['類型'] == '主控添加') |
                (filtered['類型'] == 'LIVE')
        )
        filtered = filtered[final_filter_condition].copy()
        after_final_filter = len(filtered)
        removed_no_broadcast = before_final_filter - after_final_filter

        if removed_no_broadcast > 0:
            callback(f"移除了 {removed_no_broadcast} 個完全沒有播出記錄的條目")

        # ========== 時間格式處理 ==========
        if 'CueDown' in filtered.columns:
            filtered['CueDown'] = filtered['CueDown'].astype(str).str.extract(r'(\d{2}:\d{2}:\d{2}:\d{3})',
                                                                              expand=False)
        if '實際播出' in filtered.columns:
            filtered['實際播出'] = filtered['實際播出'].astype(str).str.extract(r'(\d{2}:\d{2}:\d{2}:\d{3})',
                                                                                expand=False)

        # ========== 清理欄位 ==========
        columns_to_drop = ['持續時間 (Dura)', '表定播放時間_dt', '播出時間_dt', '字幕實際播出時間_dt', '__source__',
                           '__rptm_hour__']
        filtered = filtered.drop(columns=[col for col in columns_to_drop if col in filtered.columns], errors='ignore')

        # ========== 欄位順序 ==========
        desired_order = [
            '表定播放時間 (RPTM)', 'CueDown', '實際播出', '時間差',
            '節目名 (Name)', '段', '入點 (TIn)', '實際長度',
            '影片檔名', '類型', '字幕指令', '字幕Dur', '字幕實際播出時間',
            '事件時間', '事件類型', '事件詳情'
        ]
        # 新增事件專用欄（初始化為空白，事件行才會填入）
        for ev_col in ['事件時間', '事件類型', '事件詳情']:
            if ev_col not in filtered.columns:
                filtered[ev_col] = ''
        filtered = filtered[[col for col in desired_order if col in filtered.columns]]
        filtered = filtered.fillna('').replace('nan', '')

        if '段' in filtered.columns:
            filtered['段'] = pd.to_numeric(filtered['段'], errors='coerce').fillna(0).astype(int).astype(str)

        callback(f"最終記錄數: {len(filtered)}")

        # ========== 生成 Excel ==========
        output_file = os.path.join(
            output_path,
            f"{config['final_output_prefix']}{target_str_yyyymmdd}大洋播控AS_RUN_LOG.xlsx"
        )
        os.makedirs(os.path.dirname(output_file), exist_ok=True)

        # 1. 先用「字幕單模板」建立/更新 AS_RUN_LOG 檔
        subtitle_template = os.path.join(
            output_path,
            f"{config['final_output_prefix']}{target_str_yyyymmdd}大洋播控字幕單記錄.xlsx"
        )

        # 如果模板存在，而且 AS_RUN_LOG 尚未生成，就先從模板 copy 一份
        if os.path.exists(subtitle_template) and not os.path.exists(output_file):
            shutil.copy2(subtitle_template, output_file)
            callback(f"以字幕單模板建立 AS_RUN_LOG 檔案: {output_file}")
        elif not os.path.exists(output_file):
            callback(f"找不到字幕單模板，將直接建立新檔: {subtitle_template}", 'info')

        # 2. 載入 AS_RUN_LOG 檔，只重建 ASRUNLOG，其他頁籤（包含字幕單）一律保留
        try:
            if os.path.exists(output_file):
                wb = openpyxl.load_workbook(output_file)
                callback(f"載入後目前頁籤: {', '.join(wb.sheetnames)}")

                if "ASRUNLOG" in wb.sheetnames:
                    del wb["ASRUNLOG"]
                    callback("已刪除舊的 ASRUNLOG 頁籤")

                ws = wb.create_sheet("ASRUNLOG", 0)
            else:
                wb = openpyxl.Workbook()
                default_sheet = wb.active
                wb.remove(default_sheet)
                ws = wb.create_sheet("ASRUNLOG", 0)
        except Exception as load_error:
            callback(f"載入/建立 Excel 檔失敗，改用新檔: {load_error}", 'error')
            wb = openpyxl.Workbook()
            default_sheet = wb.active
            wb.remove(default_sheet)
            ws = wb.create_sheet("ASRUNLOG", 0)

        # 添加標題行
        ws.append(list(filtered.columns))

        # 設定第一行凍結
        ws.freeze_panes = 'A2'

        # ========== Excel 格式化 ==========
        microsoft_font = Font(name='微軟正黑體', size=11)
        header_font = Font(name='微軟正黑體', size=11, bold=True)

        light_yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        light_olive_fill = PatternFill(start_color="D4E6B7", end_color="D4E6B7", fill_type="solid")
        light_orange_fill = PatternFill(start_color="FFDAB9", end_color="FFDAB9", fill_type="solid")
        light_blue_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        white_bold_font = Font(name='微軟正黑體', size=11, bold=True, color="FFFFFF")

        highlight_cols = ['CueDown', '實際播出', '時間差', '字幕實際播出時間']

        # 設定標題行格式和字體
        for col_idx, col_name in enumerate(filtered.columns, 1):
            cell = ws[f'{get_column_letter(col_idx)}1']
            cell.font = header_font
            if col_name in highlight_cols:
                cell.fill = light_yellow_fill
            elif col_name == '事件時間':
                cell.fill = red_fill
                cell.font = white_bold_font

        # 設定欄寬（原 13 欄 + 3 個事件欄）
        column_widths = [25.43, 14.71, 14.71, 7.71, 39.43, 3.43, 13.29, 13.29, 17.86, 9.86, 13.71, 13.29, 19.57,
                         10.0, 18.71, 50.0]
        for i, width in enumerate(column_widths, 1):
            if i <= len(filtered.columns):
                ws.column_dimensions[get_column_letter(i)].width = width

        # 添加開始記錄行
        start_entry = [''] * len(filtered.columns)
        start_entry[filtered.columns.get_loc('表定播放時間 (RPTM)')] = f'{target_str_yyyymmdd} 00:00:00:00'
        start_entry[filtered.columns.get_loc('節目名 (Name)')] = '當日播出開始'
        start_entry[filtered.columns.get_loc('類型')] = 'START'
        ws.append(start_entry)

        # 設定開始記錄行的字體
        for col_idx in range(1, len(filtered.columns) + 1):
            ws[f'{get_column_letter(col_idx)}2'].font = microsoft_font

        # ========== 建立事件時間排序佇列（完全不動 filtered）==========
        # 每個事件取 時間 不含秒严格比對 RPTM
        pending_events = list(events) if events else []
        ev_ptr = 0  # 下一個尚未插入的事件索引

        def _write_event_row(ev_dict, ws, n_cols, microsoft_font):
            """ 事件行按字幕格式寫入：清空 ASRUNLOG 主欄，填入專用事件欄，差異底色 """
            col_names = list(filtered.columns)
            ev_row = [''] * n_cols

            # 類型欄：標示為事件
            if '類型' in col_names:
                ev_row[col_names.index('類型')] = '事件'
            # 事件時間欄
            if '事件時間' in col_names:
                ev_row[col_names.index('事件時間')] = ev_dict['時間']
            # 事件類型欄
            if '事件類型' in col_names:
                ev_row[col_names.index('事件類型')] = ev_dict['事件類型']
            # 事件詳情欄（清理後 + 機器/角色標示）
            if '事件詳情' in col_names:
                cleaned = _clean_ev_info(ev_dict['事件詳情'])
                ev_row[col_names.index('事件詳情')] = (
                    f"{cleaned}  [{ev_dict['機器']} {ev_dict['角色']}]"
                )

            ws.append(ev_row)
            cur_row = ws.max_row
            ev_fill = PatternFill(start_color=ev_dict['ev_color'],
                                  end_color=ev_dict['ev_color'], fill_type='solid')
            # 一般欄：正黑 11pt 非粗黑
            ev_font_normal = Font(name='微軟正黑體', size=11, bold=False)
            # 詳情欄：正黑 9pt 非粗黑
            ev_font_detail = Font(name='微軟正黑體', size=9, bold=False)

            detail_col_idx = col_names.index('事件詳情') + 1 if '事件詳情' in col_names else None
            for ci in range(1, n_cols + 1):
                ws[f'{get_column_letter(ci)}{cur_row}'].fill = ev_fill
                if detail_col_idx and ci == detail_col_idx:
                    ws[f'{get_column_letter(ci)}{cur_row}'].font = ev_font_detail
                else:
                    ws[f'{get_column_letter(ci)}{cur_row}'].font = ev_font_normal


        # 添加數據行並設定字體和顏色
        n_cols = len(filtered.columns)
        # 取得 RPTM 欄索引，用于時間比對
        rptm_col_idx = (
            filtered.columns.get_loc('表定播放時間 (RPTM)')
            if '表定播放時間 (RPTM)' in filtered.columns else None
        )
        for row_idx, row in enumerate(filtered.itertuples(index=False), start=3):

            # === 在寫入此行前，先檢查是否有事件該插入 ===
            if pending_events and rptm_col_idx is not None:
                row_time_str = str(list(row)[rptm_col_idx])
                # 如果該行為空(NaN)，則不與事件比對(保持 row_time_str_clean 為空)，直到遇到下一個有時間的行
                # 或者我們可以使用一個極大的時間來強制在這個循環中跳過事件插入
                # 但為了準確，如果 row_time_str 無效，我們不應在此行前插入未到時間的事件
                row_time_str_clean = row_time_str[9:17] if len(row_time_str) >= 17 else ''
                
                if row_time_str_clean:
                    while ev_ptr < len(pending_events):
                        ev = pending_events[ev_ptr]
                        ev_hms = (ev['_ts_dt'].strftime('%H:%M:%S') if ev['_ts_dt'] else '99:99:99')
                        if ev_hms <= row_time_str_clean:
                            _write_event_row(ev, ws, n_cols, microsoft_font)
                            ev_ptr += 1
                        else:
                            break

            ws.append(list(row))
            row_idx = ws.max_row  # 取得實際寫入的行號

            # 為每個儲存格設定字體
            for col_idx in range(1, len(filtered.columns) + 1):
                cell = ws[f'{get_column_letter(col_idx)}{row_idx}']
                cell.font = microsoft_font

            type_col_idx = filtered.columns.get_loc('類型')
            if len(row) > type_col_idx and row[type_col_idx] == 'LIVE':
                for col_idx in range(1, len(filtered.columns) + 1):
                    ws[f'{get_column_letter(col_idx)}{row_idx}'].fill = light_blue_fill
            elif len(row) > type_col_idx and row[type_col_idx] == 'PROG':
                for col_idx in range(1, len(filtered.columns) + 1):
                    ws[f'{get_column_letter(col_idx)}{row_idx}'].fill = light_olive_fill
            elif len(row) > type_col_idx and row[type_col_idx] == '主控添加':
                for col_idx in range(1, len(filtered.columns) + 1):
                    ws[f'{get_column_letter(col_idx)}{row_idx}'].fill = light_orange_fill

                # 清空前9欄
                columns_to_clear = [
                    '表定播放時間 (RPTM)', 'CueDown', '實際播出', '時間差',
                    '節目名 (Name)', '段', '入點 (TIn)', '實際長度', '影片檔名'
                ]
                for col_name in columns_to_clear:
                    if col_name in filtered.columns:
                        col_idx_to_clear = filtered.columns.get_loc(col_name) + 1
                        ws.cell(row=row_idx, column=col_idx_to_clear, value="")
            elif len(row) > type_col_idx and row[type_col_idx] == '字幕':
                for col_idx in range(1, len(filtered.columns) + 1):
                    ws[f'{get_column_letter(col_idx)}{row_idx}'].fill = light_yellow_fill

                columns_to_clear = [
                    '表定播放時間 (RPTM)', 'CueDown', '實際播出', '時間差',
                    '節目名 (Name)', '段', '入點 (TIn)', '實際長度', '影片檔名'
                ]
                for col_name in columns_to_clear:
                    if col_name in filtered.columns:
                        col_idx_to_clear = filtered.columns.get_loc(col_name) + 1
                        ws.cell(row=row_idx, column=col_idx_to_clear, value="")
            else:
                for col_idx, col_name in enumerate(filtered.columns, 1):
                    if col_name in highlight_cols:
                        ws[f'{get_column_letter(col_idx)}{row_idx}'].fill = light_yellow_fill

        # ========== 補寫剩餘事件 (重要：解決日終事件遺漏) ==========
        while pending_events and ev_ptr < len(pending_events):
            _write_event_row(pending_events[ev_ptr], ws, n_cols, microsoft_font)
            ev_ptr += 1

        # ========== 儲存最終報表 ==========
        try:
            callback(f"存檔前頁籤: {', '.join(wb.sheetnames)}")  # ← 新增這行
            wb.save(output_file)
            current_time = datetime.now().strftime('%H:%M:%S')
            callback(f"✓ Excel檔案儲存成功 [{current_time}]: {output_file}", 'success')
            callback(f"   保留的頁籤: {', '.join(wb.sheetnames)}")

            if os.path.exists(output_file):
                file_size = os.path.getsize(output_file)
                callback(f"檔案大小: {file_size:,} bytes")

                try:
                    verification_df = pd.read_excel(output_file, sheet_name='ASRUNLOG')
                    callback(f"ASRUNLOG 頁籤驗證成功，包含 {len(verification_df)} 筆記錄")
                except Exception as verify_error:
                    callback(f"檔案驗證失敗: {verify_error}", 'error')
            else:
                callback(f"⚠ 警告：檔案儲存後未找到: {output_file}", 'error')

        except PermissionError:
            timestamp = int(time.time())
            backup_path = output_file.replace('.xlsx', f'_{timestamp}.xlsx')
            wb.save(backup_path)
            current_time = datetime.now().strftime('%H:%M:%S')
            callback(f"原檔案被佔用，已儲存至備用檔案 [{current_time}]: {backup_path}", 'info')
            output_file = backup_path
        except Exception as e:
            callback(f"儲存最終報表失敗: {e}", 'error')
            return

        # ========== 複製到網路路徑 ==========
        try:
            network_path = r"\\172.25.1.218\工程部\主控及訊號中心\PRES\大洋ASRUN LOG"
            network_output_file = os.path.join(network_path,
                                               f"{config['final_output_prefix']}{target_str_yyyymmdd}大洋播控AS_RUN_LOG.xlsx")
            os.makedirs(network_path, exist_ok=True)

            if not os.access(network_path, os.W_OK):
                callback(f"網路路徑無寫入權限: {network_path}", 'error')
            else:
                shutil.copy2(output_file, network_output_file)
                callback(f"最終報表已同步至網路路徑: {network_output_file}", 'success')
        except PermissionError:
            timestamp = int(time.time())
            network_backup_file = network_output_file.replace('.xlsx', f'_{timestamp}.xlsx')
            shutil.copy2(output_file, network_backup_file)
            callback(f"網路路徑檔案被佔用，已複製至備用檔案: {network_backup_file}", 'info')
        except Exception as e:
            callback(f"同步最終報表至網路路徑失敗: {e}", 'error')

    except Exception as e:
        callback(f"generate_final_excel_complete 錯誤: {e}", 'error')
        import traceback
        callback(f"詳細錯誤: {traceback.format_exc()}", 'error')


# ==================== 主入口 ====================
def run_full_process_logic(channel: str, target_date: datetime, paths_config: Dict,
                           output_path: str, callback=None):
    """
    主處理流程入口

    Args:
        channel: 頻道代碼 ('CH02' 或 'CH04')
        target_date: 目標日期
        paths_config: 路徑配置字典，可包含以下鍵值:
            - ch02_lis: CH02 LIS 檔案路徑
            - ch02_log: CH02 網路日誌路徑
            - ch02_subtitle: CH02 字幕日誌路徑
            - ch04_lis: CH04 LIS 檔案路徑
            - ch04_log: CH04 網路日誌路徑
            - ch04_subtitle: CH04 字幕日誌路徑
        output_path: 輸出路徑
        callback: 回調函數，用於記錄訊息 callback(message, level='info')
    """
    if callback is None:
        callback = lambda msg, level='info': print(f"[{level.upper()}] {msg}")

    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    channel_display_map = {'CH02': '壹電影', 'CH04': '年代MUCH', 'CH05': '壹綜合'}
    channel_display = channel_display_map.get(channel, channel)
    date_str = target_date.strftime('%Y年%m月%d日')

    callback("=" * 60)
    callback(f"[{timestamp}] 開始 {channel_display} {date_str} 的完整處理流程")
    callback("=" * 60)

    try:
        # 取得頻道配置
        config = get_channel_config(channel, target_date)

        # 套用自訂路徑（如果有提供）
        channel_lower = channel.lower()
        if f'{channel_lower}_lis' in paths_config:
            config['lis_base_path'] = paths_config[f'{channel_lower}_lis']
        if f'{channel_lower}_log' in paths_config:
            config['log_base_path'] = paths_config[f'{channel_lower}_log']
        if f'{channel_lower}_subtitle' in paths_config:
            config['subtitle_base_path'] = paths_config[f'{channel_lower}_subtitle']

        # === 確保網路連線 (針對 CH05 151/152 等) ===
        if channel in ["CH04", "CH05", "CH02"]:
            import subprocess
            for base_path in config.get('log_base_paths', []):
                m_ip = re.search(r'\\\\([\d.]+)\\', base_path)
                if m_ip:
                    ip = m_ip.group(1)
                    target = f"\\\\{ip}\\c$"
                    callback(f"正在嘗試建立連線: {target}...")
                    # 嘗試以預設認證連線 (Abc1234)
                    try:
                        subprocess.run(f'net use "{target}" "Abc1234" /user:"Administrator" /persistent:no', 
                                       shell=True, capture_output=True, timeout=10)
                    except:
                        pass

        # 步驟 1: 處理 LIS 檔案
        callback(f"步驟 1: 處理 {channel_display} {date_str} 的 LIS 檔案...")
        lis_df, lis_items = process_lis_file_complete(config, target_date, callback)

        if lis_df.empty:
            callback("✗ LIS 檔案處理失敗", 'error')
            return False

        callback(f"✓ LIS 檔案處理完成，處理了 {len(lis_df)} 筆記錄", 'success')

        # 步驟 2: 處理網路日誌
        callback(f"步驟 2: 處理 {channel_display} {date_str} 的網路日誌...")
        log_df = process_network_log_complete(config, target_date, callback)

        if log_df is not None:
            callback(f"✓ 網路日誌處理完成", 'success')
        else:
            callback(f"⚠ 網路日誌處理完成，但沒有找到記錄", 'info')

        # 步驟 3: 處理字幕日誌
        callback(f"步驟 3: 處理 {channel_display} {date_str} 的字幕日誌...")
        subtitle_df = process_subtitle_log_complete(config, target_date, callback)

        if subtitle_df is not None:
            callback(f"✓ 字幕日誌處理完成，找到 {len(subtitle_df)} 筆字幕記錄", 'success')
        else:
            callback(f"⚠ 字幕日誌處理完成，但沒有找到記錄", 'info')

        # 步驟 3.5: 解析 DxAplayerLog 重點事件
        callback(f"步驟 3.5: 解析 {channel_display} {date_str} 的 DxAplayerLog 重點事件...")
        try:
            events = process_dxaplayer_events(config, target_date, callback)
            if events:
                callback(f"✓ DxAplayerLog 重點事件: {len(events)} 筆", 'success')
            else:
                callback("ℹ DxAplayerLog 未找到重點事件", 'info')
        except Exception as ev_err:
            callback(f"⚠ 重點事件解析失敗（不影響主報表）: {ev_err}", 'error')
            events = []

        # 步驟 4: 合併所有資料並生成 Excel
        callback(f"步驟 4: 合併 {channel_display} {date_str} 的所有資料...")
        generate_final_excel_complete(lis_df, log_df, subtitle_df, config, target_date, output_path, callback,
                                      events=events)

        callback("=" * 60)
        callback(f"🎉 {channel_display} {date_str} 完整處理流程完成！", 'success')
        callback("=" * 60)

        return True

    except Exception as e:
        callback("=" * 60)
        callback(f"✗ 處理過程中發生錯誤: {str(e)}", 'error')
        callback("=" * 60)
        import traceback
        callback(f"詳細錯誤: {traceback.format_exc()}", 'error')
        return False


# ==================== 使用範例 ====================
if __name__ == "__main__":
    # 配置範例
    paths_config = {
        'ch02_lis': r'\\172.25.80.121\c$\dayang\obin_24051401_2.8\autobak',
        'ch02_log': r'\\172.25.80.121\c$\dayang\obin_24051401_2.8\log',
        'ch02_subtitle': r'\\172.25.80.32\d$\emagic3_5_patch_2025-4-41_test\Other',
        'ch04_lis': r'\\172.25.80.141\c$\dayang\obin_24051401_2.8\autobak',
        'ch04_log': r'\\172.25.80.141\c$\dayang\obin_24051401_2.8\log',
        'ch04_subtitle': r'\\172.25.80.34\d$\emagic3_5_patch_2025-4-41_test\Other',
    }

    output_path = r'C:\Users\jerry.lee\PycharmProjects\pythonProject'
    target_date = datetime.now()

    # 執行 CH02 處理
    print("\n執行 CH02 處理...")
    success = run_full_process_logic('CH02', target_date, paths_config, output_path)

    if success:
        print("\n✓ CH02 處理成功完成")
    else:
        print("\n✗ CH02 處理失敗")

    # 執行 CH04 處理
    print("\n執行 CH04 處理...")
    success = run_full_process_logic('CH04', target_date, paths_config, output_path)

    if success:
        print("\n✓ CH04 處理成功完成")
    else:
        print("\n✗ CH04 處理失敗")